import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Files
excel_path = "Rekap_Perusahaan_2021_2024.xlsx"
log_path = "idx_downloader.log"
output_excel = "Laporan_Kegagalan_Unduh.xlsx"

# 1. Load company names map from Excel
try:
    df_companies = pd.read_excel(excel_path)
    df_companies.columns = [c.strip() for c in df_companies.columns]
    company_map = dict(zip(
        df_companies['Kode'].astype(str).str.strip(), 
        df_companies['Nama Perusahaan'].astype(str).str.strip()
    ))
except Exception as e:
    print(f"Warning: Gagal membaca {excel_path} untuk nama perusahaan: {e}")
    company_map = {}

# Helper to categorize file name
def categorize_file(filename: str) -> str:
    fn_lower = filename.lower()
    annual_indicators = ["annualreport", "annual report", "annual_report", "laporan tahunan", "laporan_tahunan"]
    for ind in annual_indicators:
        if ind in fn_lower:
            return "Laporan Tahunan"
    if fn_lower.startswith("ar20") or fn_lower.startswith("ar_20") or fn_lower.startswith("ar-20") or fn_lower.startswith("ar 20"):
        return "Laporan Tahunan"
    financial_indicators = ["financialstatement", "financial statement", "financial_statement", "laporan keuangan", "laporan_keuangan", "lkb"]
    for ind in financial_indicators:
        if ind in fn_lower:
            return "Laporan Keuangan"
    if fn_lower.startswith("lk ") or fn_lower.startswith("lk_") or fn_lower.startswith("lk-") or fn_lower.startswith("lk20"):
        return "Laporan Keuangan"
    if "lk" in fn_lower or "financial" in fn_lower:
        return "Laporan Keuangan"
    elif "annual" in fn_lower or "report" in fn_lower or "ar" in fn_lower:
        return "Laporan Tahunan"
    return "Dokumen Lainnya"

# 2. Parse log
failures = []
current_ticker = None
current_year = None
current_company = None

# Regex patterns
proc_pattern = re.compile(r"Memproses\s+(\w+)\s+\((.*?)\)\s+-\s+Tahun\s+(\d{4}).*")
warn_pattern = re.compile(r"Tidak ada file PDF ditemukan untuk\s+(\w+)\s+(\d{4})")
err_pattern = re.compile(r"Gagal mengunduh file\s+(.*?):\s+(.*)")

if not os.path.exists(log_path):
    print(f"Error: {log_path} tidak ditemukan!")
    exit(1)

with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
    lines = f.readlines()

# Find the index of the last start
start_idx = 0
for idx, line in enumerate(lines):
    if "==== IDX Bulk Document Downloader Started ====" in line:
        start_idx = idx

print(f"Membaca log dari baris {start_idx + 1} sampai {len(lines)}")

for line in lines[start_idx:]:
    # Check for processing line
    proc_match = proc_pattern.search(line)
    if proc_match:
        current_ticker = proc_match.group(1).strip()
        current_company = proc_match.group(2).strip()
        current_year = int(proc_match.group(3))
        continue
        
    # Check for warning line (no PDF found in API)
    warn_match = warn_pattern.search(line)
    if warn_match:
        ticker = warn_match.group(1).strip()
        year = int(warn_match.group(2))
        company = company_map.get(ticker, ticker)
        failures.append({
            "Kode Emiten": ticker,
            "Nama Emiten": company,
            "Tahun": year,
            "Jenis Dokumen": "Laporan Keuangan/Tahunan",
            "Nama File": "-",
            "Alasan Kegagalan": "Tidak ada file PDF ditemukan di IDX API"
        })
        continue
        
    # Check for error line (download failed)
    err_match = err_pattern.search(line)
    if err_match:
        filename = err_match.group(1).strip()
        reason = err_match.group(2).strip()
        category = categorize_file(filename)
        failures.append({
            "Kode Emiten": current_ticker,
            "Nama Emiten": current_company if current_company else company_map.get(current_ticker, current_ticker),
            "Tahun": current_year,
            "Jenis Dokumen": category,
            "Nama File": filename,
            "Alasan Kegagalan": f"Download gagal: {reason}"
        })

# 3. Create DataFrame
if not failures:
    print("Tidak ditemukan data kegagalan dalam log terakhir.")
    df_failures = pd.DataFrame(columns=["Kode Emiten", "Nama Emiten", "Tahun", "Jenis Dokumen", "Nama File", "Alasan Kegagalan"])
else:
    df_failures = pd.DataFrame(failures)
    # Deduplicate in case of duplicate logs
    df_failures = df_failures.drop_duplicates()

print(f"Total kegagalan unik yang diekstrak: {len(df_failures)}")

# Sort by Ticker and Year
if not df_failures.empty:
    df_failures = df_failures.sort_values(by=["Kode Emiten", "Tahun"])

# 4. Save to Excel with professional styles
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kegagalan Unduhan"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Add title block
ws.merge_cells("A1:F1")
ws['A1'] = "LAPORAN KEGAGALAN UNDUHAN DOKUMEN IDX"
ws['A1'].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Add metadata block
ws['A2'] = "Tanggal Pembuatan:"
ws['B2'] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
ws['A3'] = "Sumber Log:"
ws['B3'] = "idx_downloader.log (Sesi Terakhir)"
for r in [2, 3]:
    ws[f'A{r}'].font = Font(name="Calibri", size=10, bold=True)
    ws[f'B{r}'].font = Font(name="Calibri", size=10, italic=True)

ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 20

# Add headers
headers = ["Kode Emiten", "Nama Emiten", "Tahun", "Jenis Dokumen", "Nama File", "Alasan Kegagalan"]
ws.row_dimensions[5].height = 28

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_idx)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
row_start = 6
font_data = Font(name="Calibri", size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for r_idx, row_data in enumerate(df_failures.values, row_start):
    ws.row_dimensions[r_idx].height = 20
    is_even = (r_idx % 2 == 0)
    current_fill = fill_even if is_even else fill_odd
    
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.font = font_data
        cell.fill = current_fill
        cell.border = thin_border
        
        # Formatting
        if c_idx in [1, 3]: # Kode, Tahun
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < 5: # Skip title and metadata for width calculation
            continue
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    # Add a bit of padding
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Specific custom widths for better presentation
ws.column_dimensions['A'].width = 15  # Kode Emiten
ws.column_dimensions['B'].width = 40  # Nama Emiten
ws.column_dimensions['C'].width = 12  # Tahun
ws.column_dimensions['D'].width = 22  # Jenis Dokumen
ws.column_dimensions['E'].width = 40  # Nama File
ws.column_dimensions['F'].width = 50  # Alasan Kegagalan

# Save workbook
wb.save(output_excel)
print(f"Laporan berhasil dibuat di: {os.path.abspath(output_excel)}")
